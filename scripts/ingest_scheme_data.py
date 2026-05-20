"""
Scheme data ingest — builds the scheme_detail table in Supabase.

Sources:
  1. new additions/CSAT_Scheme_Summaries_v3.xlsx
       Imis_id = centre_scheme_id (confirmed: all 5,968 IDs match centre_scheme_id in master)
       Scheme_summary = AI narrative
       Key_issues     = semicolon-separated tags

  2. new additions/Copy of 1.scheme_master_data_all 21 Nov.xlsx  (Sheet 1)
       centre_scheme_id → state_scheme_id (IMIS ID), scheme_name, district,
       division, blocks, sub_divisions, planned_fhtc, achieved_fhtc

Steps:
  1. Run create_scheme_detail_table.sql in Supabase SQL editor FIRST.
  2. Then run:  python scripts/ingest_scheme_data.py

The 47 edge cases (one state IMIS ID → multiple centre IDs) are stored
as separate rows — each centre_scheme_id is unique, IMIS ID may repeat.
"""
import sys, json, time
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import requests

# ── Supabase config ───────────────────────────────────────────────────────────
SUPABASE_URL = "https://ubdgohqafxdugonrchkv.supabase.co"
SERVICE_KEY  = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InViZGdvaHFhZnhkdWdvbnJjaGt2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzMyODM0NCwiZXhwIjoyMDkyOTA0MzQ0fQ"
    ".a2Glerx0gX_cAhFtsrLYtyRLmLhGIwg8ayYWBSWZ4Q8"
)
HEADERS = {
    "apikey":        SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "resolution=merge-duplicates,return=minimal",
}
TABLE_URL   = f"{SUPABASE_URL}/rest/v1/scheme_detail"
BATCH_SIZE  = 200

BASE = r"C:\Users\Vibha Beria\OneDrive\Desktop\jjm dashboard\new additions"
SUMMARIES_FILE = BASE + r"\CSAT_Scheme_Summaries_v3.xlsx"
MASTER_FILE    = BASE + r"\Copy of 1.scheme_master_data_all 21 Nov.xlsx"


# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(v):
    """Return a clean string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ('NULL', 'None', '') else s

def to_int(v):
    """Float/string → int, or None."""
    try:
        return int(float(str(v)))
    except (TypeError, ValueError):
        return None


# ── Step 1: Load CSAT summaries ───────────────────────────────────────────────
print("Reading CSAT summaries…")
summaries: dict[int, dict] = {}
wb = openpyxl.load_workbook(SUMMARIES_FILE, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:
    cid = to_int(row[0])
    if cid is None:
        continue
    summaries[cid] = {
        'scheme_summary': clean(row[1]),
        'key_issues':     clean(row[2]),
    }
wb.close()
print(f"  Loaded {len(summaries):,} summaries (keyed by centre_scheme_id)")


# ── Step 2: Load & aggregate scheme master ────────────────────────────────────
print("Reading scheme master data…")

# Aggregation dict — keyed by centre_scheme_id
master: dict[int, dict] = {}

wb2 = openpyxl.load_workbook(MASTER_FILE, data_only=True)
ws2 = wb2['Sheet 1 - districts(45)']
rows2 = list(ws2.iter_rows(values_only=True))
wb2.close()

# Headers: district(0) division(1) scheme_name(2) centre_scheme_id(3)
#          state_scheme_id(4) sub_divisions(5) blocks(6) panchayat_name(7)
#          village_name(8) planned_fhtc(9) achieved_fhtc(10)
#          latitude(11) longitude(12) so_name(13) ...

skipped = 0
for row in rows2[1:]:
    cid = to_int(row[3])
    if cid is None:
        skipped += 1
        continue

    if cid not in master:
        master[cid] = {
            'imis_id':      to_int(row[4]),
            'scheme_name':  clean(row[2]),
            'district':     clean(row[0]),
            'division':     clean(row[1]),
            'blocks':       set(),
            'sub_divisions':set(),
            'planned_fhtc': to_int(row[9]),
            'achieved_fhtc':to_int(row[10]),
        }

    # Accumulate blocks and sub-divisions (schemes often span multiple)
    b = clean(row[6])
    if b:
        master[cid]['blocks'].add(b)
    sd = clean(row[5])
    if sd:
        master[cid]['sub_divisions'].add(sd)

print(f"  Aggregated {len(master):,} unique centre_scheme_ids ({skipped} rows skipped — no centre ID)")

# Resolve edge cases: IMIS IDs that appear in multiple centre_scheme_ids
from collections import Counter
imis_counter = Counter(v['imis_id'] for v in master.values() if v['imis_id'])
multi_imis = {k for k, c in imis_counter.items() if c > 1}
print(f"  Edge cases: {len(multi_imis)} IMIS IDs map to multiple centre_scheme_ids "
      f"(stored as separate rows — centre_scheme_id remains the unique key)")


# ── Step 3: Build combined records ────────────────────────────────────────────
print("Merging summaries + master data…")
records = []

for cid in sorted(master.keys()):
    m = master[cid]
    s = summaries.get(cid, {})

    records.append({
        'centre_scheme_id': cid,
        'imis_id':          m['imis_id'],
        'scheme_name':      m['scheme_name'],
        'district':         m['district'],
        'division':         m['division'],
        'blocks':           ', '.join(sorted(m['blocks'])) or None,
        'sub_divisions':    ', '.join(sorted(m['sub_divisions'])) or None,
        'planned_fhtc':     m['planned_fhtc'],
        'achieved_fhtc':    m['achieved_fhtc'],
        'scheme_summary':   s.get('scheme_summary'),
        'key_issues':       s.get('key_issues'),
    })

with_summary    = sum(1 for r in records if r['scheme_summary'])
without_summary = len(records) - with_summary
print(f"  Total records: {len(records):,}")
print(f"  With CSAT summary: {with_summary:,}")
print(f"  Without summary:   {without_summary:,} (master-only schemes)")


# ── Step 4: Upload in batches ─────────────────────────────────────────────────
print(f"\nUploading {len(records):,} records in batches of {BATCH_SIZE}…")
errors = 0
for i in range(0, len(records), BATCH_SIZE):
    batch = records[i : i + BATCH_SIZE]
    resp  = requests.post(TABLE_URL, headers=HEADERS, data=json.dumps(batch, default=str))
    if resp.status_code not in (200, 201, 204):
        print(f"  ERROR at batch {i//BATCH_SIZE + 1}: {resp.status_code} — {resp.text[:200]}")
        errors += 1
    else:
        done = min(i + BATCH_SIZE, len(records))
        print(f"  Uploaded {done:>6,} / {len(records):,}", end='\r')
    time.sleep(0.05)   # stay within rate limits

print(f"\nDone — {len(records) - errors*BATCH_SIZE:,} records uploaded, {errors} batch errors")
if errors == 0:
    print("✓  scheme_detail table is ready for the dashboard.")
else:
    print("⚠  Some batches failed — check errors above and re-run if needed.")
